"""
National Highways' public "7-day closure report" XLSX (advance notice of
FULL closures, published before VSS signs are activated -- so it can show
closures days before they'd appear via the API, which only reports what's
currently signed on the road). See:
https://nationalhighways.co.uk/roads-and-travel/live-travel-updates/road-closure-report/

IMPORTANT: this file's exact column layout HAS since been verified against
a real download (see the header-row match for "Road number"/"Direction"/
"Location"/"Scheduled start time"/"Scheduled end time"/"Closure details,
including diversions" -- confirmed correct). The parser below still
matches column headers flexibly by keyword and logs exactly what it finds
in every sheet in case that ever changes, so if a future report restructures
things, the build log will show the real headers to fix against, and the
site still builds fine from the other source(s) in the meantime (a parsing
failure here is a warning, not a fatal error).

CONFIRMED LIVE BUG, now fixed: the XLSX's own download URL is NOT stable.
National Highways mints a new hashed media path (e.g.
/media/qsnnq4d0/7-day-closure-report.xlsx) each time they republish this
report -- observed directly: a URL captured once kept returning a valid,
successfully-parsing XLSX with a plausible row count indefinitely, it just
quietly stopped being the CURRENT file, silently missing rows the live
report page's current link did have. This wasn't visible via the source's
🟢/🟡/🔴 status either, since the stale URL still returns 200 with real
content -- "fetch succeeded" and "fetch got the CURRENT data" turned out
to be different things. Fixed by discovering the real, current XLSX link
from the report page's own HTML on every build (see discover_xlsx_url()),
rather than trusting a URL captured once and hardcoded -- the same
"scrape the live page, don't trust a fixed data URL" reasoning
sources/traffic_scotland.py already uses, for the same underlying reason.
"""
from __future__ import annotations

import io
import re
import urllib.error
import urllib.parse
import urllib.request
from datetime import date as date_cls
from datetime import datetime
from zoneinfo import ZoneInfo

from sources import status

XLSX_FILENAME_RE = re.compile(r"7-day-closure-report\.xlsx$", re.IGNORECASE)


def discover_xlsx_url(report_page_url: str) -> str | None:
    """Scrape the CURRENT XLSX download link off the closure report page's
    own HTML, since National Highways rotates the file's hashed media path
    each time they republish it (see module docstring -- confirmed live,
    not hypothetical). Matches by filename ("...7-day-closure-report.xlsx"),
    not by hash, so it survives that rotation. Returns None (never raises)
    if the page can't be fetched, beautifulsoup4 isn't installed, or no
    matching link is found -- any of which the caller treats as "discovery
    failed", falling back to a last-known URL if one was configured."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("Warning: beautifulsoup4 is not installed -- can't discover the "
              "current XLSX link from the report page.")
        return None

    print(f"Fetching {report_page_url} to discover the current XLSX link ...")
    req = urllib.request.Request(report_page_url, headers={"User-Agent": "route-closures-build/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as e:
        print(f"Warning: HTTP {e.code} {e.reason} fetching the closure report page "
              f"({report_page_url}) -- can't discover the current XLSX link.")
        return None
    except Exception as e:  # noqa: BLE001 -- discovery failing isn't fatal, see caller
        print(f"Warning: failed to fetch the closure report page ({e}) -- "
              f"can't discover the current XLSX link.")
        return None

    soup = BeautifulSoup(html, "html.parser")
    link = soup.find("a", href=XLSX_FILENAME_RE)
    if link is None or not link.get("href"):
        print("Warning: couldn't find a '...7-day-closure-report.xlsx' link on the "
              "report page -- its structure may have changed.")
        return None

    resolved = urllib.parse.urljoin(report_page_url, link["href"])
    print(f"  discovered current XLSX link: {resolved}")
    return resolved

XLSX_HEADER_SYNONYMS: dict[str, set[str]] = {
    "road_name": {"road", "roadname", "route", "roadnumber"},
    "direction": {"direction"},
    "location_description": {
        "location", "locationdescription", "section", "extent",
        "closuresection", "closurelocation", "between", "workslocation",
    },
    "start_datetime": {
        "startdate", "start", "closurestartdate", "datefrom",
        "closurestart", "startdatetime", "scheduledstarttime", "scheduledstart",
    },
    "end_datetime": {
        "enddate", "end", "closureenddate", "dateto",
        "closureend", "enddatetime", "scheduledendtime", "scheduledend",
    },
    "comment": {
        "description", "comment", "comments", "details",
        "workdescription", "scheme", "schemedescription", "reason",
        "closuredetailsincludingdiversions", "closuredetails",
    },
    "validity_status": {"status", "closurestatus"},
}

# How many leading rows to scan (per sheet) looking for the header row --
# some sheets have a title row above the real column headers.
MAX_HEADER_SCAN_ROWS = 6

# Real data only ever uses 6 columns (Road number, Direction, Location,
# Scheduled start/end time, Closure details), but a sheet can report a
# much wider "used" range than that -- openpyxl reports it out to
# wherever cell formatting was ever applied, even to cells nobody put
# data in (seen on a live sheet: hundreds of extra reported columns with
# no content). Reading all of those for every row is real overhead, not
# just noisy logging, so this bounds how many columns are actually read.
# Generous headroom beyond the known 6 in case a future column is added.
MAX_COLUMNS_TO_READ = 20


def normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def to_iso_datetime(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date_cls)):
        return value.isoformat()
    return str(value).strip()


def compute_validity_status(start_iso: str, end_iso: str, now: datetime) -> str:
    """CONFIRMED LIVE BUG, now fixed: this report has no "status" column of
    its own (see the real header dump this project checked against --
    Road number/Direction/Location/Scheduled start/end/Closure details,
    no status field at all), so every row here used to be hardcoded to
    "planned" permanently -- including hours after a closure had actually
    started. A user directly confirmed this: an M6 J39-40 closure showed
    "Planned" on the site at 02:29 while genuinely inside its own
    22:00-06:00 window, with National Highways' own app showing the
    corresponding entry as "Active" at the same moment.

    Fixed the same way sources/traffic_scotland.py already solves the
    identical problem for its own data (duplicated here rather than
    imported, matching this project's pattern of no cross-imports
    between source modules -- see e.g. national_highways_traffic_search.py's
    clean_iso for the same reasoning): 'active' only while `now` genuinely falls
    within [start, end], 'planned' otherwise. If `end` is missing/
    unparseable but `start` isn't, compare against `start` alone (active
    once begun). Falls back to 'planned' only when even `start` is
    missing/unparseable, since there's nothing real left to compare
    against -- this report is exclusively advance NOTICE, so an entry
    with no parseable start time was never going to be safely assumed
    active regardless."""
    start = None
    end = None
    if start_iso:
        try:
            start = datetime.fromisoformat(start_iso)
        except ValueError:
            start = None
    if end_iso:
        try:
            end = datetime.fromisoformat(end_iso)
        except ValueError:
            end = None

    if start and end:
        return "active" if start <= now <= end else "planned"
    if start:
        return "active" if now >= start else "planned"
    return "planned"


def find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """Scan the first few rows of a sheet for the one that looks like a
    header row (i.e. has a cell matching a known road-name synonym), since
    the real header row isn't always row 1 -- some sheets have a title row
    above it. Returns (row_index, col_map) or (None, None) if not found."""
    for row_idx, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        headers = [normalize_header(h) for h in row]
        col_map: dict[str, int] = {}
        for field, synonyms in XLSX_HEADER_SYNONYMS.items():
            for idx, h in enumerate(headers):
                if h in synonyms:
                    col_map[field] = idx
                    break
        if "road_name" in col_map:
            return row_idx, col_map
    return None, None


def trim_trailing_empty(row: tuple) -> tuple:
    """Drop trailing None/empty cells for cleaner logging. Some sheets
    report far more "used" columns than actually contain data -- openpyxl
    reports a sheet's used range out to wherever cell formatting was ever
    applied, even to cells that were never filled in, so iter_rows() can
    return a row padded with hundreds of Nones past the real content.
    This only affects what gets printed, not parsing (which still uses
    the full row and looks up columns by index via col_map)."""
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return tuple(row)


def fetch_from_xlsx_advance_notice(
    report_page_url: str | None = None,
    fallback_xlsx_url: str | None = None,
    url: str | None = None,
) -> list[dict]:
    """Fetches the CURRENT XLSX advance-notice report and parses it.

    report_page_url (recommended): the human-facing report page --
    discover_xlsx_url() scrapes the real, current download link from it
    on every call, so this never goes stale even as National Highways
    rotates the file's hashed media path (see module docstring).

    fallback_xlsx_url (optional): used only if discovery from
    report_page_url fails (page unreachable, structure changed, etc.) --
    a safety net, not the primary path. Since National Highways rotates
    this URL, whatever you put here WILL eventually go stale too; it's
    there so a temporary discovery failure doesn't mean losing this
    source's data for that build, not as a long-term substitute for
    discovery.

    url (deprecated): the OLD calling convention -- a single hardcoded
    XLSX URL with no discovery at all. Still supported so existing
    routes.yaml configs don't break outright, but this is exactly the
    pattern that caused a real, confirmed-live staleness bug (see module
    docstring) -- migrate to report_page_url instead.
    """
    label = "Advance Notice (XLSX)"

    if report_page_url:
        xlsx_url = discover_xlsx_url(report_page_url)
        if xlsx_url is None:
            if fallback_xlsx_url:
                print(f"  falling back to the last-known XLSX URL: {fallback_xlsx_url}")
                xlsx_url = fallback_xlsx_url
            else:
                status.record_status(label, ok=False, error="could not discover the current XLSX link and no fallback_xlsx_url was configured")
                return []
    elif url:
        xlsx_url = url
    else:
        status.record_status(label, ok=False, error="neither report_page_url nor url was configured")
        return []

    try:
        import openpyxl
    except ImportError:
        print("Warning: openpyxl is not installed -- skipping the XLSX "
              "advance-notice source (add it to requirements.txt).")
        status.record_status(label, ok=False, error="openpyxl not installed")
        return []

    print(f"Fetching {xlsx_url} ...")
    req = urllib.request.Request(xlsx_url, headers={"User-Agent": "route-closures-build/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read()
    except urllib.error.HTTPError as e:
        print(f"Warning: XLSX fetch failed with HTTP {e.code} {e.reason} -- skipping this source.")
        status.record_status(label, ok=False, error=f"HTTP {e.code} {e.reason}")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 -- best-effort source, never fatal
        print(f"Warning: could not open the XLSX file ({e}) -- skipping this source.")
        status.record_status(label, ok=False, error=f"could not open file: {e}")
        return []

    closures: list[dict] = []
    row_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_col=MAX_COLUMNS_TO_READ, values_only=True))
        if not rows:
            continue  # empty sheet

        header_idx, col_map = find_header_row(rows)
        if header_idx is None:
            preview = trim_trailing_empty(rows[0]) if rows else ()
            print(f"  sheet '{sheet_name}': no recognizable 'road' column in the "
                  f"first {MAX_HEADER_SCAN_ROWS} rows -- skipping this sheet "
                  f"(first row: {preview}).")
            continue

        print(f"  sheet '{sheet_name}' headers (row {header_idx + 1}): "
              f"{trim_trailing_empty(rows[header_idx])}")

        def get(row, field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        sheet_rows = 0
        now = datetime.now(ZoneInfo("Europe/London")).replace(tzinfo=None)
        for row in rows[header_idx + 1:]:
            road_name = get(row, "road_name")
            if not road_name:
                continue
            row_counter += 1
            sheet_rows += 1
            start_iso = to_iso_datetime(get(row, "start_datetime"))
            end_iso = to_iso_datetime(get(row, "end_datetime"))
            # This report has no "status" column of its own (see real
            # headers this project confirmed against) -- always computed
            # from the real start/end window now, never a hardcoded
            # "planned" (see compute_validity_status()'s docstring for
            # the confirmed-live bug this replaced).
            explicit_status = str(get(row, "validity_status") or "").strip().lower()
            validity_status = explicit_status or compute_validity_status(start_iso, end_iso, now)
            closures.append({
                "record_id": f"xlsx-{sheet_name}-{row_counter}",
                "road_name": str(road_name).strip(),
                "direction": str(get(row, "direction") or "").strip(),
                "location_description": str(get(row, "location_description") or "").strip(),
                "comment": str(get(row, "comment") or "").strip(),
                "start_datetime": start_iso,
                "end_datetime": end_iso,
                "validity_status": validity_status,
                "cause_type": "advanceNoticeFullClosure",
                "lanes_restricted": None,
                "lanes_operational": 0,  # this report is full closures only
                "source_label": "Advance notice (full closure)",
            })
        print(f"  sheet '{sheet_name}': parsed {sheet_rows} closure rows")

    print(f"Parsed {len(closures)} total rows from the XLSX advance-notice report")
    status.record_status(label, ok=True, count=len(closures))
    return closures
