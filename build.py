#!/usr/bin/env python3
"""
Build a static site showing roadworks/closures for user-defined routes.

A "route" (e.g. Axis, Omega) has a northbound and southbound direction,
and each direction is a chain of one or more "legs" -- road sections
travelled in order (e.g. M6 J45-26, then M58, then M57 J6-4).

Data can come from either:
  - the National Highways Road & Lane Closures API v2 (live, needs an
    API key), which returns a nested DATEX II JSON payload -- this is
    flattened by normalize_datex_response() into the same simple record
    shape the rest of this script works with, or
  - a pre-flattened JSON mirror (e.g. a GitHub-hosted snapshot) using
    the same field names already.

Reads routes.yaml for route definitions and source config, fetches
closures, filters records per leg, and renders static HTML into
./_site, ready to be published as a GitHub Pages artifact.

Usage:
    python build.py

Environment:
    NATIONAL_HIGHWAYS_API_KEY   required when routes.yaml site.source
                                 is "national_highways_api"
"""
from __future__ import annotations

import io
import json
import os
import re
import shutil
import urllib.error
import urllib.request
from datetime import date as date_cls
from datetime import datetime, timedelta, timezone
from pathlib import Path

import yaml
from jinja2 import Environment, FileSystemLoader

ROOT = Path(__file__).parent
TEMPLATES_DIR = ROOT / "templates"
STATIC_DIR = ROOT / "static"
OUTPUT_DIR = ROOT / "_site"

NATIONAL_HIGHWAYS_DEFAULT_BASE_URL = "https://api.data.nationalhighways.co.uk"

# Junction numbers show up in free text like "between J35 and J36", "M62
# eastbound Jct 36 to Jct 37", or spelled out as "Junction 40" / "Junctions
# 40 to 39" (seen in the National Highways advance-notice XLSX) -- the
# second number in a spelled-out range often has no "J"/"Junction" prefix
# of its own, so an optional trailing "to/and/-<number>" is captured too.
JUNCTION_RE = re.compile(
    r'\b(?:Junction|Junc|Jct|J)s?\.?\s*(\d+)(?:\s*(?:to|and|-|\u2013)\s*(\d+))?',
    re.IGNORECASE,
)

# road_name is sometimes blank; fall back to parsing it off the front of
# the free-text comment, e.g. "M62 eastbound Jct 36...".
ROAD_RE = re.compile(r'^(M\d+[A-Z]?|A\d+\(M\)|A\d+)\b')


def load_routes() -> dict:
    with open(ROOT / "routes.yaml") as f:
        return yaml.safe_load(f)


# ---------------------------------------------------------------------
# Fetching + normalizing closures from either data source
# ---------------------------------------------------------------------

def fetch_json(url: str, headers: dict | None = None) -> tuple[dict, dict]:
    """Returns (payload, response_headers) so callers can inspect pagination
    signals like a Link header, which json.load() alone would discard."""
    if url.startswith("file://"):
        return json.load(open(url[len("file://"):])), {}
    req = urllib.request.Request(url, headers=headers or {})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            payload = json.load(resp)
            return payload, dict(resp.headers)
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise SystemExit(
            f"HTTP {e.code} {e.reason} calling:\n  {url}\n\n"
            f"Response body from the API:\n{body}\n"
        ) from None


def find_next_page_url(payload: dict, response_headers: dict) -> str | None:
    """
    Best-effort detection of a "next page" link, since National Highways'
    docs don't publish the exact pagination mechanics for this API (their
    changelog only notes a "pagination refinement" patch in May 2025).
    Checks the two most common conventions: an RFC 5988 Link header, and a
    handful of common JSON field names for a continuation URL. If neither
    is present, this returns None and the caller assumes a single page.
    """
    link_header = response_headers.get("Link") or response_headers.get("link")
    if link_header:
        for part in link_header.split(","):
            segments = part.split(";")
            if len(segments) >= 2 and 'rel="next"' in segments[1]:
                return segments[0].strip().strip("<>")

    for key in ("nextLink", "@odata.nextLink", "next", "nextPageLink", "nextPage"):
        value = payload.get(key)
        if isinstance(value, str) and value:
            return value

    return None


def normalize_datex_response(payload: dict) -> list[dict]:
    """
    Flatten a DATEX II JSON response from the National Highways API into
    a list of closure-location dicts using the same field names used
    throughout this script (road_name, direction, location_description,
    comment, start_datetime, end_datetime, validity_status, cause_type,
    lanes_restricted, lanes_operational, record_id) -- one dict per
    physical location segment. Multiple segments can share a record_id
    (they belong to the same real-world closure); rows_for_leg() collapses
    those back down to one row per closure when rendering.
    """
    flat: list[dict] = []
    situations = payload.get("D2Payload", {}).get("situation", [])

    for situation in situations:
        for sr in situation.get("situationRecord", []):
            # The v2 API wraps the actual record under a type-specific key
            # (e.g. "sitRoadOrCarriagewayOrLaneManagement"); unwrap it if
            # present. If a future response is already unwrapped, sr itself
            # already has the fields we need, so fall back to sr.
            record = sr
            for wrapper_key in (
                "sitRoadOrCarriagewayOrLaneManagement",
                "sitRoadOrCarriagewayOrLaneManagementExtensionG",
            ):
                if wrapper_key in sr and isinstance(sr[wrapper_key], dict):
                    record = sr[wrapper_key]
                    break

            record_id = record.get("idG") or record.get("id") or ""
            validity = record.get("validity", {})
            time_spec = validity.get("validityTimeSpecification", {})
            cause = record.get("cause", {})
            comments = record.get("generalPublicComment", [])
            comment = comments[0].get("comment", "") if comments else ""

            shared = {
                "record_id": record_id,
                "validity_status": validity.get("validityStatus", "unknown"),
                "start_datetime": time_spec.get("overallStartTime", ""),
                "end_datetime": time_spec.get("overallEndTime", ""),
                "cause_type": cause.get("causeType", ""),
                "comment": comment,
            }

            groups = (
                record.get("locationReference", {})
                      .get("locLocationGroupByList", {})
                      .get("locationContainedInGroup", [])
            )

            if not groups:
                # No location breakdown available -- still emit one row
                # so the closure isn't silently dropped from the feed.
                flat.append({
                    **shared,
                    "road_name": "", "direction": "", "location_description": "",
                    "lanes_restricted": None, "lanes_operational": None,
                })
                continue

            for entry in groups:
                lin_loc = entry.get("locLinearLocation", {})
                supp = lin_loc.get("supplementaryPositionalDescription", {})
                single = entry.get("locSingleRoadLinearLocation", {})
                within_list = single.get("linearWithinLinearElement", [])
                within = within_list[0] if within_list else {}
                linear_elem = within.get("linearElement", {}).get("locLinearElementByCode", {})

                carriageways = supp.get("carriageway", [])
                lanes_restricted = lanes_operational = None
                if carriageways:
                    impact = carriageways[0].get("carriagewayExtensionG", {}).get("impactOnCarriageway", {})
                    lanes_restricted = impact.get("numberOfLanesRestricted")
                    lanes_operational = impact.get("numberOfOperationalLanes")

                flat.append({
                    **shared,
                    "road_name": linear_elem.get("roadName", ""),
                    "direction": within.get("directionOnLinearSection", ""),
                    "location_description": supp.get("locationDescription", ""),
                    "lanes_restricted": lanes_restricted,
                    "lanes_operational": lanes_operational,
                })

    return flat


def fetch_from_national_highways_api(site_cfg: dict) -> list[dict]:
    api_key = os.environ.get("NATIONAL_HIGHWAYS_API_KEY", "").strip()
    if not api_key:
        raise SystemExit(
            "NATIONAL_HIGHWAYS_API_KEY is not set.\n"
            "  - In GitHub Actions: add it as a repo secret "
            "(Settings -> Secrets and variables -> Actions) and pass it to\n"
            "    the build step's `env:` block -- see the README.\n"
            "  - Locally: export NATIONAL_HIGHWAYS_API_KEY=your-key-here "
            "before running build.py."
        )

    base_url = site_cfg.get("api_base_url", NATIONAL_HIGHWAYS_DEFAULT_BASE_URL)
    closure_type = site_cfg.get("closure_type")  # "planned" / "unplanned" / None (both)
    lookahead_days = site_cfg.get("lookahead_days", 29)
    if lookahead_days > 29:
        print(f"Note: lookahead_days ({lookahead_days}) is at or above the API's "
              f"30-day maximum window; clamping to 29 to leave a safety margin.")
        lookahead_days = 29

    now = datetime.now(timezone.utc)
    start = now.strftime("%Y-%m-%dT%H:%M:%S")
    end = (now + timedelta(days=lookahead_days)).strftime("%Y-%m-%dT%H:%M:%S")

    params = [f"startDateTime={start}", f"endDateTime={end}"]
    if closure_type:
        params.append(f"closureType={closure_type}")
    url = f"{base_url.rstrip('/')}/roads/v2.0/closures?{'&'.join(params)}"

    headers = {
        "Ocp-Apim-Subscription-Key": api_key,
        "X-Response-MediaType": "application/json",
        "X-Data-Format": "DATEXII",
        "Accept": "application/json",
        "User-Agent": "route-closures-build/1.0",
    }

    closures: list[dict] = []
    page_num = 1
    max_pages = 50  # safety cap so a pagination bug can't loop forever
    while url and page_num <= max_pages:
        print(f"Fetching {url} ...")
        payload, response_headers = fetch_json(url, headers=headers)
        page_closures = normalize_datex_response(payload)
        closures.extend(page_closures)
        print(f"  page {page_num}: {len(page_closures)} closure-location records")

        next_url = find_next_page_url(payload, response_headers)
        if next_url and next_url != url:
            url = next_url
            page_num += 1
        else:
            url = None

    if page_num > 1:
        print(f"Followed {page_num} page(s), {len(closures)} total closure-location records")
    else:
        print(f"Loaded {len(closures)} closure-location records from the live API "
              f"(single page -- no pagination link found in the response)")

    if not closures:
        print("Warning: 0 records parsed. If this is unexpected, the API's "
              "response shape may differ from what this script expects -- "
              "check the raw payload keys.")
    return closures


def fetch_from_flat_mirror(site_cfg: dict) -> tuple[list[dict], str]:
    url = site_cfg["data_url"]
    print(f"Fetching {url} ...")
    data, _headers = fetch_json(url, headers={"User-Agent": "route-closures-build/1.0"})
    closures = data["closures"]
    for c in closures:
        c.setdefault("record_id", c.get("idG") or c.get("id"))
    print(f"Loaded {len(closures)} closure records (feed updated {data.get('updated')})")
    feed_updated = format_dt(data.get("updated", ""))
    return closures, feed_updated


def load_closures(site_cfg: dict) -> tuple[list[dict], str]:
    """Returns (closures, feed_updated_label). feed_updated_label is "" when
    the source has no separate "last updated" timestamp of its own (e.g.
    the live API, which is fetched fresh on every build)."""
    source = site_cfg.get("source", "flat_json")
    if source == "national_highways_api":
        closures = fetch_from_national_highways_api(site_cfg)
        for c in closures:
            c.setdefault("source_label", "Live API")
        return closures, ""
    closures, feed_updated = fetch_from_flat_mirror(site_cfg)
    for c in closures:
        c.setdefault("source_label", "Feed")
    return closures, feed_updated


# ---------------------------------------------------------------------
# Additional source: National Highways' public "7-day closure report"
# XLSX (advance notice of FULL closures, published before VSS signs are
# activated -- so it can show closures days before they'd appear via the
# API, which only reports what's currently signed on the road). See:
# https://nationalhighways.co.uk/roads-and-travel/live-travel-updates/road-closure-report/
#
# IMPORTANT: this file's exact column layout has not been verified against
# a real download (no network access was available to inspect it while
# writing this). The parser below matches column headers flexibly by
# keyword and logs exactly what it finds in every sheet, so if the guesses
# below are wrong, the build log will show the real headers to fix them
# against, and the site still builds fine from the other source(s) in the
# meantime (a parsing failure here is a warning, not a fatal error).
# ---------------------------------------------------------------------

XLSX_HEADER_SYNONYMS: dict[str, set[str]] = {
    "road_name": {"road", "roadname", "route", "roadnumber"},
    "direction": {"direction"},
    "location_description": {
        "location", "locationdescription", "section", "extent",
        "closuresection", "closurelocation", "between", "workslocation",
    },
    "start_datetime": {
        "startdate", "start", "closurestartdate", "datefrom",
        "closurestart", "startdatetime", "scheduledstarttime", "scheduledstart",
    },
    "end_datetime": {
        "enddate", "end", "closureenddate", "dateto",
        "closureend", "enddatetime", "scheduledendtime", "scheduledend",
    },
    "comment": {
        "description", "comment", "comments", "details",
        "workdescription", "scheme", "schemedescription", "reason",
        "closuredetailsincludingdiversions", "closuredetails",
    },
    "validity_status": {"status", "closurestatus"},
}

# How many leading rows to scan (per sheet) looking for the header row --
# some sheets have a title row above the real column headers.
MAX_HEADER_SCAN_ROWS = 6


def normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def to_iso_datetime(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date_cls)):
        return value.isoformat()
    return str(value).strip()


def find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """Scan the first few rows of a sheet for the one that looks like a
    header row (i.e. has a cell matching a known road-name synonym), since
    the real header row isn't always row 1 -- some sheets have a title row
    above it. Returns (row_index, col_map) or (None, None) if not found."""
    for row_idx, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        headers = [normalize_header(h) for h in row]
        col_map: dict[str, int] = {}
        for field, synonyms in XLSX_HEADER_SYNONYMS.items():
            for idx, h in enumerate(headers):
                if h in synonyms:
                    col_map[field] = idx
                    break
        if "road_name" in col_map:
            return row_idx, col_map
    return None, None


def fetch_from_xlsx_advance_notice(url: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("Warning: openpyxl is not installed -- skipping the XLSX "
              "advance-notice source (add it to requirements.txt).")
        return []

    print(f"Fetching {url} ...")
    req = urllib.request.Request(url, headers={"User-Agent": "route-closures-build/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read()
    except urllib.error.HTTPError as e:
        print(f"Warning: XLSX fetch failed with HTTP {e.code} {e.reason} -- skipping this source.")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 -- best-effort source, never fatal
        print(f"Warning: could not open the XLSX file ({e}) -- skipping this source.")
        return []

    closures: list[dict] = []
    row_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue  # empty sheet

        header_idx, col_map = find_header_row(rows)
        if header_idx is None:
            preview = rows[0] if rows else ()
            print(f"  sheet '{sheet_name}': no recognizable 'road' column in the "
                  f"first {MAX_HEADER_SCAN_ROWS} rows -- skipping this sheet "
                  f"(first row: {preview}).")
            continue

        print(f"  sheet '{sheet_name}' headers (row {header_idx + 1}): {rows[header_idx]}")

        def get(row, field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        sheet_rows = 0
        for row in rows[header_idx + 1:]:
            road_name = get(row, "road_name")
            if not road_name:
                continue
            row_counter += 1
            sheet_rows += 1
            closures.append({
                "record_id": f"xlsx-{sheet_name}-{row_counter}",
                "road_name": str(road_name).strip(),
                "direction": str(get(row, "direction") or "").strip(),
                "location_description": str(get(row, "location_description") or "").strip(),
                "comment": str(get(row, "comment") or "").strip(),
                "start_datetime": to_iso_datetime(get(row, "start_datetime")),
                "end_datetime": to_iso_datetime(get(row, "end_datetime")),
                "validity_status": str(get(row, "validity_status") or "planned").strip().lower() or "planned",
                "cause_type": "advanceNoticeFullClosure",
                "lanes_restricted": None,
                "lanes_operational": 0,  # this report is full closures only
                "source_label": "Advance notice (full closure)",
            })
        print(f"  sheet '{sheet_name}': parsed {sheet_rows} closure rows")

    print(f"Parsed {len(closures)} total rows from the XLSX advance-notice report")
    return closures


# ---------------------------------------------------------------------
# Additional source: Traffic Scotland (M74 / A74(M) scraper)
#
# Traffic Scotland has no simple self-service API (their real-time feeds
# require an approved-subscriber application). This scrapes their public,
# server-rendered pages in two stages:
#
#   1. Two listing pages -- /traffic-information/roadworks (current) and
#      /traffic-information/planned-roadworks (planned) -- list every
#      roadwork on Scotland's entire trunk road network. Each entry's
#      title is itself a link (e.g. "M74 J8 - J9 SB - Lane Closures") to
#      a detail page. Stage 1 just finds every such link whose title
#      mentions our target road (checking the M74/A74(M) alias) and
#      records its title + URL -- no field parsing happens here.
#
#   2. Each matched entry's own detail page
#      (/more-details?sid=...&type=roadworks) has clean, structured
#      fields: Location (often has an explicit junction range, e.g. "M74
#      J8 - J9 SB"), Direction, Starting, Ending, and a Roadwork
#      description (Works: / Traffic Management: / sometimes Diversion
#      Information:). This is a much better data source than the listing
#      page alone -- critically, it's the only place a real end date is
#      available (the listing pages only ever showed a start time).
#
# This was built and unit-tested against a real detail page fetched live
# (see test fixtures), so the field-scanning logic is trustworthy. Stage
# 1's link-discovery logic (finding title links on the listing page) is
# less certain -- it wasn't checked against the live listing page's exact
# markup while writing this. If a live run logs "found 0 ... entries" on
# a listing page, that's the signal to check scotland_find_road_links()
# against the page's real structure; a fetch failure on any individual
# detail page is logged and skipped rather than failing the whole build.
# ---------------------------------------------------------------------

TRAFFIC_SCOTLAND_ROADWORKS_URL = "https://www.traffic.gov.scot/traffic-information/roadworks"
TRAFFIC_SCOTLAND_PLANNED_ROADWORKS_URL = "https://www.traffic.gov.scot/traffic-information/planned-roadworks"
TRAFFIC_SCOTLAND_BASE_URL = "https://www.traffic.gov.scot"

# Roads that share a physical carriageway under different names/eras --
# matching any alias is treated as a match for the canonical name.
SCOTLAND_ROAD_ALIASES: dict[str, set[str]] = {
    "M74": {"M74", "A74(M)"},
}

SCOTLAND_ROAD_TOKEN_RE = re.compile(r'\b(M\d+[A-Z]?|A\d+\(M\)|A\d+)\b')

# "20th of July 2026, 8:00pm" -- ordinal day, month name, year, 12-hour
# time. No timezone is given on the site; treated as naive local (UK) time.
SCOTLAND_DATE_RE = re.compile(
    r'(\d{1,2})\w{0,2}\s+of\s+([A-Za-z]+)\s+(\d{4}),\s*(\d{1,2}):(\d{2})\s*([ap]m)',
    re.IGNORECASE,
)
SCOTLAND_MONTHS = {name.lower(): i for i, name in enumerate(
    ["January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"], start=1)}

# Detail page field labels, in the order they appear. scotland_scan_
# labeled_fields() slices the page's text between consecutive labels
# found here, so it doesn't matter if an optional one (e.g. "Diversion
# Information:") is missing for a given entry.
SCOTLAND_DETAIL_LABELS = [
    "Location", "Direction", "Starting", "Ending",
    "Days & times affected", "Roadwork description",
    "Works:", "Traffic Management:", "Diversion Information:",
    "Did you find",
]


# The road token at the start of a location segment on the LISTING page,
# e.g. "M74 (" or "A701 (" -- different from SCOTLAND_ROAD_TOKEN_RE, which
# matches a bare road name with no trailing paren (used for the clean
# detail-page Location field, e.g. "M74 J8 - J9 SB").
SCOTLAND_LISTING_ROAD_TOKEN_RE = re.compile(r'\b(M\d+[A-Z]?|A\d+\(M\)|A\d+)\s*\(')

# Matches one full entry's labeled-text block on a LISTING page, keyed off
# stable label strings rather than markup. This is how entries are found
# and their "More details" link extracted -- the listing page's own
# fields (other than the link) are otherwise NOT used for the final
# record; the detail page (stage 2) is parsed separately for the real
# data, including the end date the listing page never has.
SCOTLAND_LISTING_ENTRY_RE = re.compile(
    r'Location:\s*(?P<location>.+?)\s*'
    r'Start time:\s*(?P<start>.+?)\s*'
    r'Description:',
    re.DOTALL,
)


def fetch_text(url: str, headers: dict | None = None) -> str:
    req = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read().decode("utf-8", errors="replace")


def parse_scottish_datetime(text: str) -> str:
    m = SCOTLAND_DATE_RE.search(text)
    if not m:
        return ""
    day, month_name, year, hour, minute, ampm = m.groups()
    month = SCOTLAND_MONTHS.get(month_name.lower())
    if not month:
        return ""
    hour = int(hour)
    if ampm.lower() == "pm" and hour != 12:
        hour += 12
    elif ampm.lower() == "am" and hour == 12:
        hour = 0
    try:
        return datetime(int(year), month, int(day), hour, int(minute)).isoformat()
    except ValueError:
        return ""


def scotland_extract_road_tokens(text: str) -> list[str]:
    return SCOTLAND_ROAD_TOKEN_RE.findall(text)


def scotland_canonical_road(token: str) -> str:
    """Map a raw road token to its canonical name via SCOTLAND_ROAD_ALIASES
    (e.g. both "M74" and "A74(M)" canonicalize to "M74"), so an entry
    mentioning both isn't mistaken for a genuine cross-road entry."""
    token_upper = token.upper()
    for canonical, aliases in SCOTLAND_ROAD_ALIASES.items():
        if token_upper in {a.upper() for a in aliases}:
            return canonical
    return token_upper


def scotland_find_road_entries(html: str, aliases: set[str]) -> list[dict]:
    """Stage 1: find each entry block on a listing page whose Location
    field mentions one of the given road aliases (e.g. M74 or A74(M)),
    returning its raw location text (for the cross-road ambiguity check
    -- see fetch_from_traffic_scotland) and its "More details" link.

    Entries are plain-text blocks with stable labels (Location:/Start
    time:/Description:...), NOT clickable titles -- only a separate
    "More details" link at the end of each block goes anywhere. This
    walks up from each such link to find its own containing block (the
    nearest ancestor whose text has both "Location:" and "Start time:"),
    so each link is correctly tied to its own entry regardless of
    document order or a malformed neighboring block. The listing page's
    other fields aren't used for the final record -- the detail page is
    parsed separately in stage 2, since it has much cleaner data,
    including the end date the listing page never has."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("Warning: beautifulsoup4 is not installed -- skipping the "
              "Traffic Scotland source (add it to requirements.txt).")
        return []

    soup = BeautifulSoup(html, "html.parser")
    results = []
    seen_hrefs = set()

    for link in soup.find_all("a", string=lambda s: s and "more details" in s.lower()):
        href = link.get("href", "")
        if not href or href in seen_hrefs:
            continue

        container = link
        block_text = ""
        for _ in range(8):  # walk up a bounded number of ancestors
            container = container.parent
            if container is None:
                break
            text = container.get_text("\n", strip=True)
            if "Location:" in text and "Start time:" in text:
                block_text = text
                break

        if not block_text:
            continue

        entry_match = SCOTLAND_LISTING_ENTRY_RE.search(block_text)
        if not entry_match:
            continue
        location_text = re.sub(r'\s+', ' ', entry_match.group("location")).strip()

        tokens = {t.upper() for t in SCOTLAND_LISTING_ROAD_TOKEN_RE.findall(location_text)}
        if not (tokens & aliases):
            continue

        seen_hrefs.add(href)
        absolute_href = href if href.startswith("http") else (
            TRAFFIC_SCOTLAND_BASE_URL + href if href.startswith("/")
            else f"{TRAFFIC_SCOTLAND_BASE_URL}/{href}"
        )
        results.append({"location_text": location_text, "href": absolute_href})

    return results

def scotland_scan_labeled_fields(text: str) -> dict[str, str]:
    """Slice text into label -> content by finding where each known label
    (in SCOTLAND_DETAIL_LABELS) first occurs, then taking everything up
    to the next label that's actually present. Robust to optional
    sections being missing (e.g. no "Diversion Information:") since it
    only slices between labels that were actually found, in the order
    they appear -- not a fixed positional template."""
    positions = []
    for label in SCOTLAND_DETAIL_LABELS:
        idx = text.find(label)
        if idx != -1:
            positions.append((idx, label))
    positions.sort()

    fields = {}
    for i, (idx, label) in enumerate(positions):
        start = idx + len(label)
        end = positions[i + 1][0] if i + 1 < len(positions) else len(text)
        fields[label] = text[start:end].strip()
    return fields


def scotland_parse_detail_page(html: str, href: str) -> dict | None:
    """Stage 2: parse one entry's detail page into a closure dict (still
    missing road_name/validity_status/source_label -- the caller fills
    those in, since this function doesn't know which listing page or
    road search found it)."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return None

    soup = BeautifulSoup(html, "html.parser")
    full_text = soup.get_text("\n", strip=True)

    # Bound the scan to just the roadwork-details card, so page chrome
    # (nav/footer) can't collide with a label name.
    start_marker = full_text.find("Roadwork details")
    text = full_text[start_marker:] if start_marker != -1 else full_text

    fields = scotland_scan_labeled_fields(text)
    location = fields.get("Location", "").strip()
    if not location:
        return None  # page structure didn't match what's expected here

    direction = fields.get("Direction", "").strip()
    start_text = fields.get("Starting", "").strip()
    end_text = fields.get("Ending", "").strip()
    works = re.sub(r'\s+', ' ', fields.get("Works:", "")).strip()
    tm = re.sub(r'\s+', ' ', fields.get("Traffic Management:", "")).strip()
    diversion = re.sub(r'\s+', ' ', fields.get("Diversion Information:", "")).strip()

    comment_parts = []
    if works:
        comment_parts.append(f"Works: {works}")
    if tm:
        comment_parts.append(f"Traffic Management: {tm}")
    if diversion:
        comment_parts.append(f"Diversion: {diversion}")
    comment = " | ".join(comment_parts)

    sid_match = re.search(r'[?&]sid=([^&]+)', href)
    record_id = f"scotland-{sid_match.group(1)}" if sid_match else f"scotland-{hash(href)}"

    return {
        "record_id": record_id,
        "location_description": location,
        "direction": direction,
        "comment": comment,
        "start_datetime": parse_scottish_datetime(start_text),
        "end_datetime": parse_scottish_datetime(end_text),
        "cause_type": works,
    }


def fetch_from_traffic_scotland(road_name: str = "M74") -> list[dict]:
    aliases = SCOTLAND_ROAD_ALIASES.get(road_name, {road_name})
    pages = [
        (TRAFFIC_SCOTLAND_ROADWORKS_URL, "active"),
        (TRAFFIC_SCOTLAND_PLANNED_ROADWORKS_URL, "planned"),
    ]

    # href -> (listing_location_text, validity_status) -- de-duplicated
    # across both listing pages in case an entry somehow appears on both
    found: dict[str, tuple[str, str]] = {}
    for url, status in pages:
        print(f"Fetching {url} ...")
        try:
            html = fetch_text(url, headers={"User-Agent": "route-closures-build/1.0"})
        except urllib.error.HTTPError as e:
            print(f"Warning: HTTP {e.code} {e.reason} fetching {url} -- skipping this page.")
            continue

        entries = scotland_find_road_entries(html, aliases)
        print(f"  found {len(entries)} {road_name}-matching entr"
              f"{'y' if len(entries) == 1 else 'ies'} on this page")
        for entry in entries:
            found.setdefault(entry["href"], (entry["location_text"], status))

    if not found:
        print("Warning: 0 matching entries found from Traffic Scotland. If "
              "this is unexpected, the listing page's real block/link "
              "structure may differ from what scotland_find_road_entries() "
              "expects.")
        return []

    print(f"Fetching detail pages for {len(found)} matched entr"
          f"{'y' if len(found) == 1 else 'ies'} ...")

    results = []
    skipped_ambiguous = 0
    fetch_failures = 0
    for href, (listing_location_text, status) in found.items():
        try:
            detail_html = fetch_text(href, headers={"User-Agent": "route-closures-build/1.0"})
        except urllib.error.HTTPError as e:
            print(f"  Warning: HTTP {e.code} fetching {href} -- skipping this entry.")
            fetch_failures += 1
            continue
        except Exception as e:  # noqa: BLE001 -- one bad entry shouldn't fail the build
            print(f"  Warning: failed to fetch {href} ({e}) -- skipping this entry.")
            fetch_failures += 1
            continue

        entry = scotland_parse_detail_page(detail_html, href)
        if not entry:
            print(f"  Warning: could not parse detail page {href} -- skipping.")
            continue

        # Cross-road ambiguity guard: if the listing's location text (the
        # original context this entry was found in) mentions more than
        # one distinct road, and the DETAIL page's own clean location
        # text has no junction number of its own, skip rather than risk
        # matching on a junction that belongs to the other road (real
        # case seen in practice: an M8/M74 closure whose diversion
        # mentioned M8's own junctions 21/23, which happened to fall
        # inside the M74 leg's range).
        tokens = {t.upper() for t in scotland_extract_road_tokens(
            f"{listing_location_text} {entry['location_description']}"
        )}
        distinct_roads = {scotland_canonical_road(t) for t in tokens}
        is_cross_road = len(distinct_roads) > 1
        if is_cross_road and not _junctions_in_text(entry["location_description"]):
            skipped_ambiguous += 1
            continue

        entry["road_name"] = road_name
        entry["validity_status"] = status
        entry["lanes_restricted"] = None
        entry["lanes_operational"] = None
        entry["source_label"] = "Traffic Scotland (scraped)"
        results.append(entry)

    if skipped_ambiguous:
        print(f"  skipped {skipped_ambiguous} entr{'y' if skipped_ambiguous == 1 else 'ies'} "
              f"mentioning another road with no explicit {road_name} junction number")
    if fetch_failures:
        print(f"  {fetch_failures} detail page fetch(es) failed and were skipped")

    print(f"Parsed {len(results)} {road_name} closures from Traffic Scotland detail pages")
    return results


def load_additional_closures(site_cfg: dict) -> list[dict]:
    extra: list[dict] = []
    for source in site_cfg.get("additional_sources", []) or []:
        source_type = source.get("type")
        if source_type == "xlsx_advance_notice":
            try:
                extra.extend(fetch_from_xlsx_advance_notice(source["url"]))
            except Exception as e:  # noqa: BLE001 -- additional sources are best-effort
                print(f"Warning: additional source '{source_type}' failed ({e}) -- "
                      f"continuing without it.")
        elif source_type == "traffic_scotland_scraper":
            try:
                extra.extend(fetch_from_traffic_scotland(
                    road_name=source.get("road_name", "M74"),
                ))
            except Exception as e:  # noqa: BLE001 -- additional sources are best-effort
                print(f"Warning: additional source '{source_type}' failed ({e}) -- "
                      f"continuing without it.")
        else:
            print(f"Warning: unknown additional source type '{source_type}' -- skipping.")
    return extra


# ---------------------------------------------------------------------
# Route matching (unchanged regardless of data source -- everything is
# normalized to the same flat record shape by this point)
# ---------------------------------------------------------------------

def resolve_road_name(closure: dict) -> str:
    if closure.get("road_name"):
        return closure["road_name"]
    comment = closure.get("comment") or ""
    m = ROAD_RE.match(comment.strip())
    return m.group(1) if m else ""


def _junctions_in_text(text: str) -> list[int]:
    nums = []
    for m in JUNCTION_RE.finditer(text):
        nums.append(int(m.group(1)))
        if m.group(2):
            nums.append(int(m.group(2)))
    return nums


def extract_junctions(closure: dict) -> list[int]:
    """
    Junction numbers for route matching/sorting come from the structured
    location text only, not the free-text comment -- the comment field can
    contain diversion route instructions (e.g. "diversion via A50, rejoin
    at J24") that mention junctions with no relation to where the closure
    actually is, which would otherwise contaminate matching and sort order.
    Only fall back to the comment when location text has no junction info.
    """
    junctions = _junctions_in_text(closure.get("location_description") or "")
    if junctions:
        return junctions
    return _junctions_in_text(closure.get("comment") or "")


def closure_matches_leg(closure: dict, road_name: str, data_direction: str,
                         j_from: int | None, j_to: int | None) -> bool:
    if resolve_road_name(closure).upper() != road_name.upper():
        return False
    if (closure.get("direction") or "").lower() != data_direction.lower():
        return False
    if j_from is None or j_to is None:
        return True  # "entire road" leg -- no junction filter
    junctions = extract_junctions(closure)
    if not junctions:
        return False
    lo, hi = sorted((j_from, j_to))
    return any(lo <= j <= hi for j in junctions)


def format_dt(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        return datetime.fromisoformat(iso_str).strftime("%d %b %Y, %H:%M")
    except ValueError:
        return iso_str


def road_badge_class(road_name: str) -> str:
    """UK signage convention: motorways are blue, primary A-roads green."""
    if road_name.upper().startswith("M") or "(M)" in road_name.upper():
        return "badge-motorway"
    return "badge-primary"


def leg_direction_sort_key(closure: dict, j_from: int | None, j_to: int | None):
    """
    Order closures along the direction of travel for a leg -- e.g. for
    M6 J45 -> J26 southbound, a closure near J45 should come before one
    near J30. Falls back to start time for closures with no junction
    number in their text (which also covers "entire road" legs, where
    junction ordering isn't meaningful).
    """
    if j_from is not None and j_to is not None:
        junctions = extract_junctions(closure)
        if junctions:
            avg = sum(junctions) / len(junctions)
            # travelling from a higher to a lower junction number -> reverse
            return (0, -avg if j_from > j_to else avg)
    return (1, closure.get("start_datetime") or "")


def rows_for_leg(closures: list[dict], road_name: str, data_direction: str,
                  j_from: int | None, j_to: int | None) -> list[dict]:
    matches = [
        c for c in closures
        if closure_matches_leg(c, road_name, data_direction, j_from, j_to)
    ]

    # A single closure can have multiple matching location segments (e.g.
    # spans several junctions within one leg's range) -- collapse those
    # back down to one row per underlying closure.
    deduped: dict[object, dict] = {}
    for c in matches:
        key = c.get("record_id") or id(c)
        if key not in deduped:
            deduped[key] = c
    matches = list(deduped.values())
    matches.sort(key=lambda c: leg_direction_sort_key(c, j_from, j_to))

    rows = []
    for c in matches:
        location_text = c.get("location_description") or c.get("comment") or "\u2014"

        # If the displayed location has no junction number of its own but
        # matching/sorting still found one (via the comment fallback in
        # extract_junctions() -- e.g. a diversion instruction like "leave
        # the motorway at J22"), surface it so there's never a gap between
        # why a row is positioned where it is and what's visible about it.
        # Worded as "near" since a diversion-derived number is an inferred
        # proxy for the closure's location, not a stated fact about it.
        if not _junctions_in_text(c.get("location_description") or ""):
            fallback_junctions = extract_junctions(c)
            if fallback_junctions:
                uniq = sorted(set(fallback_junctions))
                note = "/".join(f"J{j}" for j in uniq)
                location_text = f"{location_text} \u2014 near {note}"

        rows.append({
            "location": location_text,
            "comment": c.get("comment") or "",
            "start": format_dt(c.get("start_datetime", "")),
            "end": format_dt(c.get("end_datetime", "")),
            "start_iso": c.get("start_datetime") or "",
            "end_iso": c.get("end_datetime") or "",
            "status": (c.get("validity_status") or "unknown").lower(),
            "lanes_restricted": c.get("lanes_restricted"),
            "lanes_operational": c.get("lanes_operational"),
            "cause": (c.get("cause_type") or "").replace("Work", " work").strip(),
            "source_label": c.get("source_label") or "",
        })
    return rows


def build_direction(closures: list[dict], dir_cfg: dict) -> dict:
    """Build the leg groups (each with its own rows) for one direction."""
    leg_groups = []
    total = 0
    active_total = 0
    for leg in dir_cfg["legs"]:
        rows = rows_for_leg(
            closures,
            leg["road_name"],
            leg["data_direction"],
            leg.get("junction_from"),
            leg.get("junction_to"),
        )
        leg_groups.append({
            "road_name": leg["road_name"],
            "badge_class": road_badge_class(leg["road_name"]),
            "junction_from": leg.get("junction_from"),
            "junction_to": leg.get("junction_to"),
            "rows": rows,
            "count": len(rows),
        })
        total += len(rows)
        active_total += sum(1 for r in rows if r["status"] == "active")
    return {
        "label": dir_cfg["label"],
        "leg_groups": leg_groups,
        "total": total,
        "active_total": active_total,
    }


def main() -> None:
    config = load_routes()
    site_cfg = config["site"]

    closures, feed_updated = load_closures(site_cfg)
    closures.extend(load_additional_closures(site_cfg))
    print(f"Total closures across all sources: {len(closures)}")

    generated_at = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    if not feed_updated:
        feed_updated = generated_at

    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    OUTPUT_DIR.mkdir(parents=True)
    if STATIC_DIR.exists():
        shutil.copytree(STATIC_DIR, OUTPUT_DIR / "static")

    env = Environment(loader=FileSystemLoader(str(TEMPLATES_DIR)))
    route_cards = []  # summary data for the index page

    for route in config["routes"]:
        directions_for_index = []

        for dir_key, dir_cfg in route["directions"].items():
            built = build_direction(closures, dir_cfg)
            page_id = f"{route['id']}-{dir_key}"

            html = env.get_template("route.html").render(
                site_title=site_cfg["title"],
                route_name=route["name"],
                direction_label=built["label"],
                leg_groups=built["leg_groups"],
                generated_at=generated_at,
                feed_updated=feed_updated,
            )
            (OUTPUT_DIR / f"{page_id}.html").write_text(html, encoding="utf-8")

            directions_for_index.append({
                "page": f"{page_id}.html",
                "page_id": page_id,
                "label": built["label"],
                "count": built["total"],
                "active_count": built["active_total"],
                "date_summary": [
                    {"start": r["start_iso"], "end": r["end_iso"], "status": r["status"]}
                    for lg in built["leg_groups"] for r in lg["rows"]
                ],
                "leg_summary": [
                    {
                        "road_name": lg["road_name"],
                        "badge_class": lg["badge_class"],
                        "junction_from": lg["junction_from"],
                        "junction_to": lg["junction_to"],
                    }
                    for lg in built["leg_groups"]
                ],
            })

        route_cards.append({
            "name": route["name"],
            "directions": directions_for_index,
        })

    index_html = env.get_template("index.html").render(
        site_title=site_cfg["title"],
        route_cards=route_cards,
        generated_at=generated_at,
        feed_updated=feed_updated,
    )
    (OUTPUT_DIR / "index.html").write_text(index_html, encoding="utf-8")

    total_pages = sum(len(r["directions"]) for r in route_cards) + 1
    print(f"Built {total_pages} pages into {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
